import json
 
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
 
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
from .forms import FileUploadForm, EmployeeEditForm
from .permissions import can_upload, can_edit, can_view
from core.services.file_processor import EmployeeFileProcessor
from core.services.report_service import get_movement_report
from .models import Employee, Production, Workshop

from core.permissions import login_required_custom

def _style_excel_sheet(ws, header_row=1):
    """
    Применяет оформление к листу Excel:
    Заголовок: приглушённый фон + полужирный шрифт
    Все ячейки: тонкая рамка + выравнивание
    Автоширина колонок
    """
    thin_border = Border(
        left=Side(style='thin', color='C0C0C0'),
        right=Side(style='thin', color='C0C0C0'),
        top=Side(style='thin', color='C0C0C0'),
        bottom=Side(style='thin', color='C0C0C0'),
    )
    header_font = Font(bold=True, size=10, color='333333')
    header_fill = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid')
    body_font = Font(size=10, color='444444')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row == header_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
            else:
                cell.font = body_font
                # Первые 2 колонки — текст, остальные — по центру
                cell.alignment = left if cell.column <= 2 else center

    # Автоширина
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

def _style_total_row(ws, row_num):
    """Выделяет строку итогов: полужирный шрифт + темный фон."""
    total_font = Font(bold=True, size=10, color='222222')
    total_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    for cell in ws[row_num]:
        cell.font = total_font
        cell.fill = total_fill

def login_view(request):
    """Страница входа. После успешной авторизации редиректит по роли."""
    if request.user.is_authenticated:
        return _redirect_by_role(request.user)
 
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return _redirect_by_role(user)
        else:
            messages.error(request, 'Неверный логин или пароль')
    else:
        form = AuthenticationForm()
 
    return render(request, 'authorization/login.html', {'form': form})
 
 
def logout_view(request):
    """Выход из системы"""
    logout(request)
    return redirect('core:login')
 
 
def access_denied(request):
    """Страница 403 — недостаточно прав"""
    return render(request, 'authorization/access_denied.html', status=403)
 
 
def _redirect_by_role(user):
    """
    Редирект после авторизации в зависимости от роли:
      Суперпользователь        → core:upload_file
      Администратор            → core:upload_file
      Директор по персоналу   → core:dashboard
      Редактор                 → core:upload_file
      Просмотр                 → core:dashboard
      Без группы               → core:dashboard
    """
    if user.is_superuser:
        return redirect('core:upload_file')
 
    group_names = set(user.groups.values_list('name', flat=True))
 
    if 'Администратор' in group_names:
        return redirect('core:upload_file')
 
    if 'Редактор' in group_names:
        return redirect('core:upload_file')
 
    # Директор по персоналу, Просмотр, без группы — на отчёты
    return redirect('core:dashboard')


# Основные view

def home_redirect(request):
    """Редирект с главной: авторизованных — по роли, остальных — на вход"""
    if request.user.is_authenticated:
        return _redirect_by_role(request.user)
    return redirect('core:login')

@can_upload
def upload_file(request):
    """
    View для загрузки файла с сотрудниками
    """
    if request.method == 'POST':
        # Создаём форму с данными
        form = FileUploadForm(request.POST, request.FILES)

        if form.is_valid():
            # Получаем очищенные данные
            uploaded_file = request.FILES['file']
            try:
                # Создаём обработчик и обрабатываем файл
                processor = EmployeeFileProcessor(uploaded_file)
                result = processor.process()

                # Результаты загрузки
                parts = []
                if result['success'] > 0:
                    parts.append(f"добавлено {result['success']}")
                if result['updated'] > 0:
                    parts.append(f"обновлено {result['updated']}")
                if result['error_count'] > 0:
                    parts.append(f"пропущено с ошибками {result['error_count']}")

                if parts:
                    msg = f"Загрузка завершена за {result['duration']:.1f} сек: {', '.join(parts)}."
                    if result['error_count'] > 0:
                        messages.warning(request, msg)
                    else:
                        messages.success(request, msg)
                else:
                    messages.info(request, "Файл обработан, новых изменений нет.")

                # Сообщение при ошибке загрузки
                if result['errors']:
                    request.session['upload_errors'] = result['errors'][:100]

                return redirect('core:upload_file')
            except Exception as e:
                messages.error(request, f"Ошибка: {str(e)}")
        else:
            messages.error(request, "Ошибка в форме загрузки")
    else:
        # Создаем пустую форму
        form = FileUploadForm()

    # Получаем ошибки из сессии и удаляем их
    errors = request.session.pop('upload_errors', [])

    return render(request, 'core/upload/upload_file.html', {
        'form': form,
        'errors': errors,
        'required_columns': ['ФИО', 'Табельный номер', 'Дата приема на работу'],
        'optional_columns': [
            'Дата рождения', 'Дата увольнения', 'Производство',
            'Цех', 'Причина увольнения', 'Категория рабочего', 'Должность'
        ],
    })


@can_view
def employee_list(request):
    """
    Просмотр списка всех сотрудников

    Args:
        request: HTTP запрос

    Returns:
        render: Шаблон со списком сотрудников
    """
    employees = Employee.objects.select_related(
        'production',
        'workshop',
        'dismissal_reason',
        'employee_category',
        'position'
    )

    search        = request.GET.get('search', '').strip()
    production_id = request.GET.get('production', '')
    workshop_id   = request.GET.get('workshop', '')
    status        = request.GET.get('status', '')

    # Количество строк на странице — допустимые значения фиксированы во избежание злоупотреблений
    ALLOWED_PER_PAGE = [25, 50, 100, 200]
    try:
        per_page = int(request.GET.get('per_page', 50))
        if per_page not in ALLOWED_PER_PAGE:
            per_page = 50
    except (ValueError, TypeError):
        per_page = 50

    if search:
        employees = employees.filter(
            Q(full_name__icontains=search) | Q(employee_number__icontains=search)
        )
    if production_id:
        employees = employees.filter(production_id=production_id)
    if workshop_id:
        employees = employees.filter(workshop_id=workshop_id)
    if status == 'active':
        employees = employees.filter(dismissal_date__isnull=True)
    elif status == 'dismissed':
        employees = employees.filter(dismissal_date__isnull=False)

    employees = employees.order_by('full_name')

    # Считаем до пагинации, чтобы total_count всегда отражал результат фильтрации
    total_count = employees.count()

    paginator = Paginator(employees, per_page)
    page_obj  = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj':           page_obj,
        'total_count':        total_count,
        'productions':        Production.objects.all(),
        'workshops':          Workshop.objects.all(),
        'search':             search,
        'selected_production': production_id,
        'selected_workshop':   workshop_id,
        'selected_status':     status,
        'per_page':            per_page,
        'per_page_options':    ALLOWED_PER_PAGE,
    }

    if request.headers.get('HX-Request'):
        return render(request, 'core/employee/partials/employee_table.html', context)

    return render(request, 'core/employee/employee_list.html', context)


@can_view
def employee_detail(request, employee_number):
    """
    Карточка сотрудника — только просмотр, без возможности изменений
    Точка входа перед редактированием: список → карточка → редактирование
    """
    employee = get_object_or_404(
        Employee.objects.select_related(
            'production', 'workshop', 'position',
            'employee_category', 'dismissal_reason'
        ),
        employee_number=employee_number
    )

    return render(request, 'core/employee/employee_detail.html', {
        'employee': employee,
    })


@can_edit
def employee_edit(request, employee_number):
    employee = get_object_or_404(Employee, employee_number=employee_number)

    if request.method == 'POST':
        form = EmployeeEditForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, f"Сотрудник {employee.full_name} обновлён")
            # После сохранения возвращаемся на карточку, а не в общий список
            return redirect('core:employee_detail', employee_number=employee.employee_number)
        else:
            messages.error(request, "Проверьте правильность заполнения формы")
    else:
        form = EmployeeEditForm(instance=employee)

    return render(request, 'core/employee/employee_edit.html', {
        'form': form,
        'employee': employee,
    })


@can_edit
def employee_delete(request, employee_number):
    employee = get_object_or_404(Employee, employee_number=employee_number)

    if request.method == 'POST':
        full_name = employee.full_name
        employee.delete()
        messages.success(request, f"Сотрудник {full_name} удалён")
        return redirect('core:employee_list')

    return render(request, 'core/employee/employee_confirm_delete.html', {'employee': employee})


# ==========
#  Отчёты
# ==========

@can_view
def report_headcount(request):
    """Отчёт «Численность» — срез на текущий момент."""
    from core.services.report_service import get_headcount_report

    production_id = request.GET.get('production', '')
    workshop_id   = request.GET.get('workshop', '')

    prod_id = int(production_id) if production_id else None
    ws_id   = int(workshop_id)   if workshop_id   else None

    data = get_headcount_report(production_id=prod_id, workshop_id=ws_id)

    return render(request, 'core/reports/headcount.html', {
        'data':                data,
        'productions':         Production.objects.all(),
        'workshops':           Workshop.objects.all(),
        'selected_production': production_id,
        'selected_workshop':   workshop_id,
    })


@can_view
def report_headcount_export(request):
    """Экспорт отчёта «Численность» в Excel — без стилей, только данные."""
    from core.services.report_service import get_headcount_report

    production_id = request.GET.get('production', '')
    workshop_id   = request.GET.get('workshop', '')
    prod_id = int(production_id) if production_id else None
    ws_id   = int(workshop_id)   if workshop_id   else None

    data = get_headcount_report(production_id=prod_id, workshop_id=ws_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Численность'

    # Заголовок
    ws.append(['Производство', 'Цех', 'Всего'] + data['categories'])

    for prod in data['by_production']:
        for workshop in prod['workshops']:
            row = [prod['name'], f"Цех {workshop.number}", workshop.total]
            row += [workshop.cat_values.get(cat, 0) for cat in data['categories']]
            ws.append(row)
        row = [f"Итого {prod['name']}", '', prod['total']]
        row += [prod['categories'].get(cat, 0) for cat in data['categories']]
        ws.append(row)

    row = ['ИТОГО', '', data['grand_total']]
    row += [data['grand_categories'].get(cat, 0) for cat in data['categories']]
    ws.append(row)

    # Оформление
    _style_excel_sheet(ws)
    _style_total_row(ws, ws.max_row)
    # Строки промежуточных итогов по производствам
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
        if row[0].value and str(row[0].value).startswith('Итого'):
            _style_total_row(ws, row[0].row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    from datetime import date as _date
    from urllib.parse import quote
    today = _date.today()
    filename = f"Численность_на_{today.strftime('%d.%m.%y')}.xlsx"
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    wb.save(response)
    return response


@can_view
def report_movement(request):
    """Отчёт «Движение персонала» — динамика за год."""
    from core.services.report_service import get_movement_report
    from datetime import date as _date

    production_id = request.GET.get('production', '')
    workshop_id   = request.GET.get('workshop', '')
    year_param    = request.GET.get('year', '')

    prod_id = int(production_id) if production_id else None
    ws_id   = int(workshop_id)   if workshop_id   else None
    year    = int(year_param)    if year_param    else _date.today().year

    data = get_movement_report(production_id=prod_id, workshop_id=ws_id, year=year)

    return render(request, 'core/reports/movement.html', {
        'data':                data,
        'productions':         Production.objects.all(),
        'workshops':           Workshop.objects.all(),
        'selected_production': production_id,
        'selected_workshop':   workshop_id,
        'selected_year':       str(year),
    })


@can_view
def report_movement_export(request):
    """Экспорт отчёта «Движение персонала» в Excel — без стилей, только данные."""
    from core.services.report_service import get_movement_report
    from datetime import date as _date

    production_id = request.GET.get('production', '')
    workshop_id   = request.GET.get('workshop', '')
    year_param    = request.GET.get('year', '')
    prod_id = int(production_id) if production_id else None
    ws_id   = int(workshop_id)   if workshop_id   else None
    year    = int(year_param)    if year_param    else _date.today().year

    data = get_movement_report(production_id=prod_id, workshop_id=ws_id, year=year)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Лист 1: помесячная динамика
    ws1 = wb.create_sheet('Динамика по месяцам')
    ws1.append([f'Отчёт за {year} год'])
    ws1.append([])
    ws1.append(['Месяц', 'Принято', 'Уволено', 'Разница', 'Численность на конец'])
    for row in data['monthly_rows']:
        ws1.append([row['month'], row['hired'], row['dismissed'], row['diff'], row['headcount']])
    ws1.append(['ИТОГО', data['total_hired'], data['total_dismissed'], data['total_diff'], ''])

    # Лист 2: текучесть по цехам
    ws2 = wb.create_sheet('Текучесть по цехам')
    ws2.append([f'Отчёт за {year} год'])
    ws2.append([])
    ws2.append(['Производство', 'Цех', 'Среднесписочная', 'Принято', 'Уволено', 'Разница', 'Текучесть %'])
    for ws in data['workshop_rows']:
        ws2.append([
            ws.production.name if ws.production else '—',
            f"Цех {ws.number}",
            ws.avg_count,
            ws.hired,
            ws.dismissed,
            ws.diff,
            ws.turnover_rate,
        ])

    # Лист 3: причины увольнений
    ws3 = wb.create_sheet('Причины увольнений')
    ws3.append([f'Отчёт за {year} год'])
    ws3.append([])
    ws3.append(['Причина', 'Кол-во', '%'])
    for r in data['reasons_total']:
        ws3.append([r['dismissal_reason__name'], r['count'], r['pct']])

    # Лист 4: матрица цех × причина
    if data['matrix_rows']:
        ws4 = wb.create_sheet('Причины по цехам')
        ws4.append([f'Отчёт за {year} год'])
        ws4.append([])
        header = ['Цех', 'Производство'] + data['reason_order'] + ['Итого']
        ws4.append(header)
        for row in data['matrix_rows']:
            line = [f"Цех {row['workshop']}", row['production']]
            line += [row['reasons'].get(r, 0) for r in data['reason_order']]
            line += [row['total']]
            ws4.append(line)

    # Оформление всех листов
    for sheet in wb.worksheets:
        _style_excel_sheet(sheet)
        # Строка ИТОГО — последняя на листах 1-3
        if sheet.title in ('Динамика по месяцам', 'Причины увольнений'):
            _style_total_row(sheet, sheet.max_row)
            
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    from datetime import date as _date
    from urllib.parse import quote
    today = _date.today()
    filename = f"Движение_персонала_на_{today.strftime('%d.%m.%y')}_за_{year}.xlsx"
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    wb.save(response)
    return response

@login_required_custom
def profile_view(request):
    """
    Страница профиля пользователя.
    Показывает данные учётной записи и форму смены пароля.
    При успешной смене пароля пользователь остаётся залогинен.
    """
    from django.contrib.auth import update_session_auth_hash
    from core.models import UserProfile
 
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)
 
    error = None
 
    if request.method == 'POST':
        current_password = request.POST.get('current_password', '')
        new_password     = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')
 
        if not user.check_password(current_password):
            error = 'Неверный текущий пароль'
        elif new_password != confirm_password:
            error = 'Новые пароли не совпадают'
        elif len(new_password) < 8:
            error = 'Пароль должен содержать не менее 8 символов'
        elif new_password.isdigit():
            error = 'Пароль не должен состоять только из цифр'
        elif user.check_password(new_password):
            error = 'Новый пароль не должен совпадать с текущим'
        else:
            user.set_password(new_password)
            user.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Пароль успешно изменён')
            return redirect('core:profile')
 
    return render(request, 'core/profile.html', {
        'profile': profile,
        'error':   error,
    })

@can_view
def dashboard(request):
    """Дашборд — сводная аналитика по персоналу."""
    from core.services.report_service import get_headcount_report, get_movement_report
    from datetime import date as _date
    import json

    year_param = request.GET.get('year', '')
    year = int(year_param) if year_param else _date.today().year

    headcount = get_headcount_report()
    movement = get_movement_report(year=year)

    # Общий коэффициент текучести = уволено / среднесписочная * 100
    avg_headcount = headcount['grand_total']
    if avg_headcount:
        turnover_rate = round(movement['total_dismissed'] / avg_headcount * 100, 1)
    else:
        turnover_rate = 0

    # Топ-5 цехов по текучести (для таблицы внизу)
    top_workshops = sorted(
        movement['workshop_rows'],
        key=lambda ws: ws.turnover_rate,
        reverse=True
    )[:10]

    categories_json = json.dumps(headcount['categories'])

    return render(request, 'core/reports/dashboard.html', {
        'headcount':       headcount,
        'movement':        movement,
        'turnover_rate':   turnover_rate,
        'top_workshops':   top_workshops,
        'available_years': movement['available_years'],
        'selected_year':   str(year),
        'categories_json': categories_json,
    })
